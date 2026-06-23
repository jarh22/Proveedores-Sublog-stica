import io
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import plotly.express as px
import requests
import streamlit as st

CONTROL_SHEETS = {"MENU", "VF", "CONSOLIDADO"}
PAY_LINE_PATTERN = re.compile(r"^\s*PAGO\b", re.IGNORECASE)
TOTAL_LABELS = {
    "TOTAL EJECUTADO": "executed",
    "TOTAL DE CONTRATO": "contract_total",
    "POR EJECUTAR": "remaining",
}
DATE_HINTS = ["fecha", "radicado", "fecha radicado", "fecha_radicado"]


def get_sheet_id(sheet_url: str) -> str:
    url = str(sheet_url).strip().strip('"').strip("'")
    if re.fullmatch(r"[a-zA-Z0-9-_]+", url):
        return url

    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    if match:
        return match.group(1)

    match = re.search(r"/d/e/([a-zA-Z0-9-_]+)", url)
    if match:
        return match.group(1)

    raise ValueError("URL de Google Sheet no válida.")


def load_public_sheet_url_from_secrets() -> Optional[str]:
    if hasattr(st, "secrets") and st.secrets.get("public_gsheets_url"):
        return str(st.secrets["public_gsheets_url"]).strip()

    env_value = os.environ.get("PUBLIC_GSHEETS_URL")
    if env_value:
        return str(env_value).strip()

    secrets_path = Path.cwd() / ".streamlit" / "secrets.toml"
    if secrets_path.exists():
        content = secrets_path.read_text(encoding="utf-8")
        match = re.search(r"public_gsheets_url\s*=\s*[\"'](.+?)[\"']", content)
        if match:
            return match.group(1).strip()

    return None


def configure_page() -> None:
    st.set_page_config(
        page_title="Gestión de Contratos - Bomberos Bogotá",
        page_icon="🚒",
        layout="wide",
    )

    st.markdown(
        """
        <style>
        .stApp { background: #F4F7FC; color: #0f172a; }
        .reportview-container .main { background-color: #f8fafc; }
        .card { background: #ffffff; border: 1px solid #e2e8f0; border-radius: 18px; padding: 22px; box-shadow: 0 12px 30px rgba(15, 23, 42, 0.06); }
        .metric-card { background: #ffffff; border: 1px solid #e2e8f0; border-radius: 18px; padding: 18px; }
        .kpi-big { font-size: 2.8rem; font-weight: 700; margin: 0; color: #0f172a; }
        .kpi-label { color: #475569; font-size: 0.95rem; margin: 0 0 10px; }
        .progress-container { background: #e2e8f0; border-radius: 999px; overflow: hidden; height: 20px; }
        .progress-bar { background: linear-gradient(90deg, #fb923c, #f97316); height: 100%; color: #ffffff; text-align: right; padding-right: 12px; font-weight: 700; line-height: 20px; }
        .status-pill { display: inline-block; border-radius: 999px; padding: 6px 12px; font-size: 0.85rem; font-weight: 600; color: #0f172a; background: #e2e8f0; margin: 2px 4px 4px 0; }
        .pill-green { background: #dcfce7; color: #166534; }
        .pill-yellow { background: #fef9c3; color: #713f12; }
        .pill-red { background: #fee2e2; color: #991b1b; }
        .small-label { color: #667085; font-size: 0.9rem; }
        </style>
        """,
        unsafe_allow_html=True,
    )


def format_money(amount: float) -> str:
    try:
        amount = float(amount)
    except (TypeError, ValueError):
        return "-"

    negative = amount < 0
    amount = abs(amount)
    integer = int(amount)
    decimals = int(round((amount - integer) * 100))
    integer_part = f"{integer:,}".replace(",", ".")
    result = f"${integer_part},{decimals:02d}"
    return f"-{result}" if negative else result


def parse_currency(value: object) -> Optional[float]:
    if pd.isna(value):
        return None

    text = str(value).strip()
    if not text:
        return None

    text = text.replace("\u00a0", " ")
    text = text.replace("$", "").replace("COP", "").replace("USD", "")
    text = text.replace("–", "-").replace("—", "-")
    text = re.sub(r"[^0-9,\.-]", "", text)
    if not text:
        return None

    if text.count(",") == 1 and text.count(".") >= 1:
        text = text.replace(".", "")
        text = text.replace(",", ".")
    elif text.count(",") > 1 and text.count(".") == 0:
        text = text.replace(".", "")
        text = text.replace(",", ".")
    elif text.count(",") == 1 and text.count(".") == 0:
        text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value: object) -> pd.Timestamp:
    if pd.isna(value):
        return pd.NaT

    text = str(value).strip()
    if not text:
        return pd.NaT

    parsers = ["%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"]
    for fmt in parsers:
        try:
            return pd.to_datetime(datetime.strptime(text, fmt))
        except Exception:
            continue

    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    return parsed


def load_sheet_names(sheet_url: str) -> List[str]:
    sheet_id = get_sheet_id(sheet_url)
    xlsx_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    response = requests.get(xlsx_url, timeout=40)
    if response.status_code == 200:
        try:
            import openpyxl
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(response.content), read_only=True)
            return workbook.sheetnames
        except Exception:
            pass

    html_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/htmlview"
    response = requests.get(html_url, timeout=30)
    response.raise_for_status()
    text = response.text
    names = re.findall(r'"sheetName":"([^"]+)"', text)
    if names:
        return names

    compact = re.findall(r'tab-id="(\d+)"\s+title="([^"]+)"', text)
    if compact:
        return [title for _gid, title in compact]

    raise ValueError("No se pudieron detectar las pestañas del Google Sheet.")


def load_sheet_tab(sheet_url: str, sheet_name: str) -> pd.DataFrame:
    sheet_id = get_sheet_id(sheet_url)
    csv_url = (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&sheet={requests.utils.requote_uri(sheet_name)}"
    )
    response = requests.get(csv_url, timeout=40)
    if response.status_code != 200:
        raise ValueError(
            f"No se pudo cargar la pestaña '{sheet_name}'. HTTP {response.status_code}."
        )

    text = response.text
    if not text.strip():
        return pd.DataFrame()

    dataframe = pd.read_csv(io.StringIO(text), dtype=str, keep_default_na=False)
    return dataframe


def cleanup_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(col).strip() for col in df.columns]
    drop_cols = [col for col in df.columns if df[col].astype(str).str.strip().eq("").all()]
    return df.drop(columns=drop_cols)


def detect_amount_column(df: pd.DataFrame) -> Optional[str]:
    candidates = [col for col in df.columns if col != df.columns[0]]
    scores: Dict[str, int] = {}
    for col in candidates:
        values = df[col].astype(str).astype(object)
        scores[col] = sum(1 for value in values if parse_currency(value) is not None)

    if not scores:
        return df.columns[1] if len(df.columns) > 1 else df.columns[0]

    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else (df.columns[1] if len(df.columns) > 1 else df.columns[0])


def detect_date_column(df: pd.DataFrame) -> Optional[str]:
    for col in df.columns:
        normalized = str(col).strip().lower()
        if any(hint in normalized for hint in DATE_HINTS):
            return col

    for col in df.columns[1:]:
        values = df[col].astype(str)
        parsed = values.map(parse_date)
        if parsed.notna().sum() >= 1:
            return col

    return None


def extract_contract_payments(df: pd.DataFrame, sheet_name: str) -> Tuple[pd.DataFrame, Dict[str, object], List[str]]:
    df = cleanup_dataframe_columns(df)
    if df.empty:
        return pd.DataFrame(), {}, []

    label_col = df.columns[0]
    amount_col = detect_amount_column(df)
    date_col = detect_date_column(df)
    status_cols = [col for col in df.columns if col not in {label_col, amount_col, date_col}]

    label_values = df[label_col].astype(str).fillna("").str.strip()
    payments_mask = label_values.str.contains(PAY_LINE_PATTERN)
    payments_df = df.loc[payments_mask].copy()
    payments_df["Pago"] = payments_df[label_col].astype(str).str.strip()
    payments_df["Monto"] = payments_df[amount_col].map(parse_currency).fillna(0)
    if date_col:
        payments_df["Fecha_Radicado"] = payments_df[date_col].map(parse_date)
    else:
        payments_df["Fecha_Radicado"] = pd.NaT

    for col in status_cols:
        payments_df[col] = payments_df[col].astype(str).fillna("").str.strip()

    totals: Dict[str, object] = {}
    for _, row in df.iterrows():
        label = str(row[label_col]).strip().upper()
        if label in TOTAL_LABELS:
            totals[TOTAL_LABELS[label]] = parse_currency(row.get(amount_col, "")) or 0

    totals["executed"] = totals.get("executed", payments_df["Monto"].sum())
    totals["contract_total"] = totals.get("contract_total", None)
    totals["remaining"] = totals.get(
        "remaining",
        max((totals["contract_total"] or payments_df["Monto"].sum()) - totals["executed"], 0),
    )
    totals["payments_count"] = len(payments_df)
    totals["last_payment_date"] = (
        payments_df["Fecha_Radicado"].max()
        if not payments_df["Fecha_Radicado"].isna().all()
        else pd.NaT
    )
    totals["contract_sheet"] = sheet_name
    totals["provider"] = _extract_provider_name(sheet_name, df.columns[0])
    totals["contract_title"] = sheet_name
    totals["status_columns"] = status_cols

    payments_df = payments_df[["Pago", "Monto", "Fecha_Radicado"] + status_cols]
    return payments_df.reset_index(drop=True), totals, status_cols


def _extract_provider_name(sheet_name: str, header_title: str) -> str:
    header = str(header_title or sheet_name)
    match = re.search(r"\(([^)]+)\)", header)
    if match:
        return match.group(1).strip()
    return sheet_name


def clean_consolidado_summary(df: pd.DataFrame) -> pd.DataFrame:
    df = cleanup_dataframe_columns(df)
    if df.empty:
        return df

    rename_map = {}
    for col in df.columns:
        norm = str(col).strip().upper().replace(" ", "").replace(".", "")
        if "CONTRATO" in norm and ("#" in norm or norm.startswith("CONTRATO")):
            rename_map[col] = "Contrato"
        elif "BIEN" in norm or "SERVICIO" in norm:
            rename_map[col] = "BienServicio"
        elif "PROVEEDOR" in norm:
            rename_map[col] = "Proveedor"
        elif "INICIO" in norm and "FECHA" in norm:
            rename_map[col] = "FechaInicio"
        elif "TERMIC" in norm or "TERMIN" in norm:
            rename_map[col] = "FechaTermino"
        elif "VALORINICIAL" in norm:
            rename_map[col] = "ValorInicial"
        elif "ADICION" in norm:
            rename_map[col] = "Adicion"
        elif "TOTAL" in norm and "CONTRATO" in norm:
            rename_map[col] = "TotalContrato"
        elif "VALORDEPAGO" in norm or "VALORDEPAGOTOTAL" in norm or "VALORPAGOTOTAL" in norm:
            rename_map[col] = "PagoTotal"
        elif "SALDO" in norm:
            rename_map[col] = "Saldo"
        elif norm == "PAGOS":
            rename_map[col] = "Pagos"

    df = df.rename(columns=rename_map)
    if df.columns[0] != "Contrato":
        df = df.rename(columns={df.columns[0]: "Contrato"})

    df = df.loc[df["Contrato"].astype(str).str.strip() != ""].copy()
    df["Contrato"] = df["Contrato"].astype(str).str.strip()

    for numeric_col in ["ValorInicial", "Adicion", "TotalContrato", "PagoTotal", "Saldo"]:
        if numeric_col in df.columns:
            df[numeric_col] = df[numeric_col].map(parse_currency).fillna(0)

    return df


def build_value_cards(totals: Dict[str, object]) -> None:
    executed = float(totals.get("executed", 0) or 0)
    contract_total = float(totals.get("contract_total", 0) or 0)
    remaining = float(totals.get("remaining", 0) or 0)
    coverage = round((executed / contract_total * 100) if contract_total > 0 else 0, 2)

    c1, c2, c3 = st.columns(3)
    c1.markdown("<div class='metric-card'><p class='kpi-label'>Total Contrato</p><p class='kpi-big'>" + format_money(contract_total) + "</p></div>", unsafe_allow_html=True)
    c2.markdown("<div class='metric-card'><p class='kpi-label'>Total Ejecutado</p><p class='kpi-big'>" + format_money(executed) + "</p></div>", unsafe_allow_html=True)
    c3.markdown("<div class='metric-card'><p class='kpi-label'>Por Ejecutar</p><p class='kpi-big'>" + format_money(remaining) + "</p></div>", unsafe_allow_html=True)

    st.markdown("<div class='card' style='margin-top:20px;'>", unsafe_allow_html=True)
    st.markdown("<p class='kpi-label'>Ejecución del contrato</p>", unsafe_allow_html=True)
    bar_width = min(max(coverage, 0), 100)
    st.markdown(
        f"<div class='progress-container'><div class='progress-bar' style='width: {bar_width}%'>{bar_width}%</div></div>",
        unsafe_allow_html=True,
    )
    st.markdown("</div>", unsafe_allow_html=True)


def build_monthly_chart(payments_df: pd.DataFrame) -> Optional[px.Figure]:
    if payments_df.empty:
        return None

    chart_df = payments_df.copy()
    chart_df = chart_df[chart_df["Fecha_Radicado"].notna()].copy()
    if chart_df.empty:
        return None

    chart_df["Mes"] = chart_df["Fecha_Radicado"].dt.to_period("M").dt.to_timestamp()
    monthly = chart_df.groupby("Mes", as_index=False)["Monto"].sum().sort_values("Mes")
    fig = px.bar(
        monthly,
        x=monthly["Mes"].dt.strftime("%b %Y"),
        y="Monto",
        labels={"Monto": "Monto Pagado", "Mes": "Mes"},
        color_discrete_sequence=["#f97316"],
        height=360,
    )
    fig.update_layout(margin=dict(l=0, r=0, t=30, b=0), plot_bgcolor="#ffffff", paper_bgcolor="#ffffff")
    fig.update_yaxes(tickprefix="$", separatethousands=True)
    return fig


def build_status_chart(payments_df: pd.DataFrame, status_cols: List[str]) -> Optional[px.Figure]:
    if payments_df.empty or not status_cols:
        return None

    records = []
    for col in status_cols:
        counts = payments_df[col].fillna("SIN DATOS").astype(str).str.strip().value_counts()
        for status, count in counts.items():
            records.append({"Estado": status or "Sin valor", "Categoria": col, "Conteo": int(count)})

    if not records:
        return None

    status_df = pd.DataFrame(records)
    fig = px.bar(
        status_df,
        x="Categoria",
        y="Conteo",
        color="Estado",
        title="Estados de pago por columna",
        height=380,
    )
    fig.update_layout(margin=dict(l=0, r=0, t=40, b=0), plot_bgcolor="#ffffff", paper_bgcolor="#ffffff")
    return fig


def show_contract_dashboard(sheet_url: str, sheet_name: str) -> None:
    raw_df = load_sheet_tab(sheet_url, sheet_name)
    payments_df, totals, status_cols = extract_contract_payments(raw_df, sheet_name)

    st.markdown(f"## Contrato {sheet_name}")
    st.markdown(f"**Proveedor:** {totals.get('provider', '-')}")
    st.markdown(f"**Título de hoja:** {raw_df.columns[0] if not raw_df.empty else sheet_name}")

    build_value_cards(totals)

    chart = build_monthly_chart(payments_df)
    status_chart = build_status_chart(payments_df, status_cols)
    if chart is not None and status_chart is not None:
        c1, c2 = st.columns(2, gap="large")
        c1.plotly_chart(chart, use_container_width=True)
        c2.plotly_chart(status_chart, use_container_width=True)
    elif chart is not None:
        st.plotly_chart(chart, use_container_width=True)
    elif status_chart is not None:
        st.plotly_chart(status_chart, use_container_width=True)
    else:
        st.info("No hay datos de fecha válidos o estados disponibles para graficar.")

    if payments_df.empty:
        st.warning("No se detectaron filas de pago válidas en esta pestaña.")
    else:
        st.markdown("<div class='card' style='margin-top:20px;'>", unsafe_allow_html=True)
        st.markdown("<p class='kpi-label'>Detalle de pagos</p>", unsafe_allow_html=True)
        payments_df_display = payments_df.copy()
        payments_df_display["Monto"] = payments_df_display["Monto"].map(format_money)
        payments_df_display["Fecha_Radicado"] = payments_df_display["Fecha_Radicado"].dt.strftime("%Y-%m-%d")
        st.dataframe(payments_df_display.fillna(""), use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

        excel_bytes = to_excel_bytes(payments_df)
        st.download_button(
            label="Descargar pagos del contrato",
            data=excel_bytes,
            file_name=f"pagos_{sheet_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with st.expander("Ver hoja original completa"):
        st.dataframe(raw_df.fillna(""), use_container_width=True)


def show_consolidado_dashboard(sheet_url: str, sheet_name: str) -> None:
    raw_df = load_sheet_tab(sheet_url, sheet_name)
    df = clean_consolidado_summary(raw_df)
    if df.empty:
        st.warning("La pestaña CONSOLIDADO no tiene datos o no pudo limpiarse correctamente.")
        return

    st.markdown("## CONSOLIDADO")
    st.markdown(
        "Resumen de contratos y ejecución. Use este informe para comparar montos contractuales, pagos totales y saldo pendiente."
    )

    total_contracts = df.shape[0]
    total_value = df["TotalContrato"].sum() if "TotalContrato" in df.columns else 0
    paid_total = df["PagoTotal"].sum() if "PagoTotal" in df.columns else 0
    total_balance = df["Saldo"].sum() if "Saldo" in df.columns else 0
    execution_ratio = round((paid_total / total_value * 100) if total_value > 0 else 0, 2)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Contratos", total_contracts)
    c2.metric("Valor Total", format_money(total_value))
    c3.metric("Pagado Total", format_money(paid_total))
    c4.metric("Saldo Total", format_money(total_balance))

    st.markdown("<div class='card' style='margin-top:20px;'>", unsafe_allow_html=True)
    st.markdown("<p class='kpi-label'>Ejecución global</p>", unsafe_allow_html=True)
    st.markdown(
        f"<div class='progress-container'><div class='progress-bar' style='width: {execution_ratio}%'>{execution_ratio}%</div></div>",
        unsafe_allow_html=True,
    )
    st.markdown("</div>", unsafe_allow_html=True)

    if "TotalContrato" in df.columns and "PagoTotal" in df.columns:
        top_chart = df.nlargest(10, "TotalContrato")[['Contrato', 'TotalContrato', 'PagoTotal']].copy()
        top_chart = top_chart.melt(id_vars='Contrato', value_vars=['TotalContrato', 'PagoTotal'], var_name='Tipo', value_name='Valor')
        fig = px.bar(
            top_chart,
            x='Contrato',
            y='Valor',
            color='Tipo',
            barmode='group',
            title='Top 10 contratos por valor y pagos',
            height=420,
        )
        fig.update_layout(margin=dict(l=0, r=0, t=40, b=120), plot_bgcolor='#ffffff', paper_bgcolor='#ffffff')
        fig.update_yaxes(tickprefix='$', separatethousands=True)
        st.plotly_chart(fig, use_container_width=True)

    st.markdown("<div class='card' style='margin-top:20px;'>", unsafe_allow_html=True)
    st.dataframe(df.fillna(""), use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)

    excel_bytes = to_excel_bytes(df)
    st.download_button(
        label="Descargar informe consolidado",
        data=excel_bytes,
        file_name="consolidado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    try:
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Datos")
        return buffer.getvalue()
    except Exception:
        return df.to_csv(index=False).encode("utf-8")


def filter_contract_tabs(sheet_names: List[str]) -> List[str]:
    valid = [name for name in sheet_names if str(name).strip().upper() not in CONTROL_SHEETS]
    return sorted(valid, key=lambda value: value)


def main() -> None:
    configure_page()
    st.title("Gestión Financiera de Contratos — Bomberos de Bogotá")
    st.markdown(
        "Control centralizado de ejecución de contratos, pagos y seguimiento por contrato usando la hoja pública de Google Sheets."
    )

    sheet_url = load_public_sheet_url_from_secrets()
    if not sheet_url:
        st.error(
            "Falta configurar la URL del Google Sheet en `st.secrets['public_gsheets_url']`, en `.streamlit/secrets.toml`, o en la variable de entorno `PUBLIC_GSHEETS_URL`."
        )
        return

    with st.spinner("Detectando pestañas y cargando datos..."):
        try:
            sheet_names = load_sheet_names(sheet_url)
        except Exception as exc:
            st.error(f"No se pudieron detectar las pestañas del Google Sheet: {exc}")
            return

    if not sheet_names:
        st.error("No se encontraron pestañas en la hoja de cálculo.")
        return

    selected_tab = st.sidebar.selectbox(
        "Seleccionar pestaña",
        sheet_names,
        index=sheet_names.index("CONSOLIDADO") if "CONSOLIDADO" in sheet_names else 0,
    )
    st.sidebar.markdown("---")
    st.sidebar.write("Esta aplicación lee la hoja pública de Google Sheets y crea un tablero de control por pestaña.")

    if selected_tab == "CONSOLIDADO":
        show_consolidado_dashboard(sheet_url, selected_tab)
    else:
        show_contract_dashboard(sheet_url, selected_tab)


if __name__ == "__main__":
    main()
